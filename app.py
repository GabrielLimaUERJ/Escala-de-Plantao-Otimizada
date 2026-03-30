import io
import math
import random
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")
st.title("📅 Sistema de Escala")

# ========================
# CONSTANTES
# ========================

TIPOS_ESPECIAIS = ["Recesso", "Final de Semana", "Noturno", "Apoio"]
INTERVALO_MINIMO = 6

PRIORIDADE_DIA_SEMANA = {
    0: 1,  # segunda-feira  → prioridade 1 (maior)
    4: 2,  # sexta-feira    → prioridade 2
    2: 3,  # quarta-feira   → prioridade 3
    3: 4,  # quinta-feira   → prioridade 4
    1: 5,  # terça-feira    → prioridade 5 (menor)
}

# Apoio só ocorre em dias de plantão noturno e final de semana (nunca recesso)
APOIO_TIPOS_VALIDOS = {"noturno", "final_semana"}

# ========================
# LOAD DATA
# ========================

@st.cache_data
def carregar_dados():
    df = pd.read_csv("dados.csv")
    df.columns = df.columns.str.strip().str.upper()
    df["NOME"] = df["NOME"].astype(str).str.strip()
    df = df.dropna(how="all", subset=["FINAL DE SEMANA", "NOTURNO", "RECESSO", "APOIO"])
    df_long = df.melt(id_vars=["NOME"], var_name="tipo_plantao", value_name="ultima_data")
    df_long = df_long.dropna(subset=["ultima_data"])
    df_long["ultima_data"] = pd.to_datetime(df_long["ultima_data"], format="%d/%m/%Y", errors="coerce")
    df_long = df_long.dropna(subset=["ultima_data"])
    return df_long


@st.cache_data
def carregar_ferias():
    df = pd.read_csv("ferias.csv")
    df.columns = df.columns.str.strip().str.upper()
    df["NOME"] = df["NOME"].astype(str).str.strip()
    df_long = df.melt(id_vars=["NOME"], var_name="data", value_name="ferias")
    df_long["data"] = pd.to_datetime(df_long["data"], format="%d/%m/%y", errors="coerce")
    df_long = df_long.dropna(subset=["data"])
    df_long = df_long[df_long["ferias"] == True]
    return df_long


@st.cache_data
def calcular_criticidade():
    df = pd.read_csv("ferias.csv")
    df.columns = df.columns.str.strip().str.upper()
    df["NOME"] = df["NOME"].astype(str).str.strip()
    colunas_datas = [c for c in df.columns if c != "NOME"]
    criticidade = {}
    for _, row in df.iterrows():
        nome = row["NOME"]
        total_ferias = sum(1 for col in colunas_datas if row[col] == True)
        criticidade[nome] = total_ferias
    return criticidade


@st.cache_data
def carregar_regioes():
    df = pd.read_csv("regioes.csv")
    df.columns = df.columns.str.strip().str.upper()
    df["NOME"] = df["NOME"].astype(str).str.strip()
    df["REGIAO"] = df["REGIAO"].astype(str).str.strip()
    return df.set_index("NOME")["REGIAO"].to_dict()


@st.cache_data
def carregar_calendario(ANO):
    df_raw = pd.read_csv("calendario_raw.csv")
    eventos = []
    for _, row in df_raw.iterrows():
        if pd.notna(row.get("Feriados")):
            data = datetime.strptime(row["Feriados"] + f"/{ANO}", "%d/%m/%Y")
            eventos.append({"data": data, "tipo": "feriado"})
        if pd.notna(row.get("Recesso")):
            data = datetime.strptime(row["Recesso"] + f"/{ANO}", "%d/%m/%Y")
            eventos.append({"data": data, "tipo": "recesso"})
    return pd.DataFrame(eventos)


df_long            = carregar_dados()
df_ferias          = carregar_ferias()
mapeamento_regioes = carregar_regioes()
criticidade_ferias = calcular_criticidade()

# ========================
# INPUTS — SIDEBAR
# ========================

st.sidebar.header("Configurações Gerais")

ANO = st.sidebar.number_input("Ano",  value=2026)
MES = st.sidebar.number_input("Mês",  value=4)

qtd_recesso = st.sidebar.number_input("Recesso por dia",         0, 50, 2)
qtd_final   = st.sidebar.number_input("Final de Semana por dia", 0, 50, 2)
qtd_noturno = st.sidebar.number_input("Noturno por dia",         0, 50, 2)
qtd_apoio   = st.sidebar.number_input("Apoio por dia",           0, 50, 0)

n_iteracoes = st.sidebar.number_input(
    "Iterações para otimização",
    min_value=1, max_value=200, value=30,
    help=(
        "Quantas vezes o algoritmo roda com permutações aleatórias das listas "
        "para encontrar o melhor resultado. Mais iterações = melhor resultado, "
        "porém mais lento. Recomendado: 30–100."
    )
)

# ========================
# PESOS DO SCORE — SIDEBAR
#
# Todos os critérios são normalizados para [0, 1] antes da ponderação,
# eliminando distorções causadas por meses com mais ou menos dias especiais.
# Os sliders abaixo controlam a importância relativa de cada critério.
#
# C1 – Cobertura das listas pré-geradas
#      slots_especiais_alocados / slots_especiais_esperados
#      Garante que o ciclo de rotação dos plantões especiais seja respeitado.
#
# C2 – Ausência de vagas vazias
#      slots_preenchidos_total / slots_esperados_total  (especiais + diurnos)
#      Penaliza qualquer dia com slot incompleto.
#
# C3 – Equilíbrio por pessoa
#      1 − coeficiente_de_variação(plantões_por_pessoa)
#      Quanto menor a variação entre pessoas, maior a nota. CV=0 → nota 1.
#
# C4 – Cobertura de datas críticas
#      diurnos_em_dias_prioritários / máximo_possível
#      Prioriza segundas, sextas e quartas (dias de mais mandados).
#
# C5 – Equilíbrio de oficiais por dia
#      1 − coeficiente_de_variação(plantões_por_dia)
#      Garante cobertura operacional mínima em emergências.
# ========================

st.sidebar.header("⚖️ Pesos do Score (ajuste por mês)")
st.sidebar.caption(
    "Os pesos são relativos entre si — não precisam somar 100. "
    "Todos os critérios são normalizados para [0, 1] antes de serem ponderados, "
    "então o score final é robusto a variações mensais na quantidade de dias especiais."
)

peso_cobertura_listas = st.sidebar.slider(
    "P1 – Cobertura das listas especiais", 0, 100, 35,
    help="% dos slots especiais preenchidos a partir da fila pré-gerada. "
         "Valor alto garante que o ciclo de rotação seja respeitado."
)
peso_vagas_vazias = st.sidebar.slider(
    "P2 – Ausência de vagas vazias", 0, 100, 25,
    help="Penaliza dias com slots incompletos em qualquer tipo de plantão."
)
peso_equil_pessoa = st.sidebar.slider(
    "P3 – Equilíbrio por pessoa", 0, 100, 20,
    help="Quanto menor a variação de plantões entre pessoas, maior a nota."
)
peso_datas_criticas = st.sidebar.slider(
    "P4 – Cobertura de datas críticas", 0, 100, 15,
    help="Prioriza o preenchimento de plantões diurnos em segundas, sextas e quartas."
)
peso_equil_dia = st.sidebar.slider(
    "P5 – Equilíbrio de oficiais por dia", 0, 100, 5,
    help="Quanto menor a variação de plantões por dia, maior a nota."
)

# ========================
# CONFIG REGIÕES — SIDEBAR
# ========================

regioes = sorted(set(mapeamento_regioes.values()))
st.sidebar.header("⚙️ Configuração por Região")

config_regioes = {}
for reg in regioes:
    with st.sidebar.expander(f"Região: {reg}"):
        qtd_diarios = st.number_input(
            f"Qtd plantões diurnos por dia - {reg}",
            min_value=0, max_value=50, value=2, key=f"diarios_{reg}"
        )
        max_mes = st.number_input(
            f"Máx plantões diurnos por pessoa no mês - {reg}",
            min_value=0, max_value=50, value=2, key=f"max_mes_{reg}"
        )
        config_regioes[reg] = {"diurnos_por_dia": qtd_diarios, "max_mes": max_mes}

# ========================
# FUNÇÕES AUXILIARES
# ========================

def esta_de_ferias(nome, data):
    return not df_ferias[
        (df_ferias["NOME"] == nome) & (df_ferias["data"] == data)
    ].empty


def tipo_do_dia(data, df_calendario):
    linha = df_calendario[df_calendario["data"] == data]
    if not linha.empty:
        tipo = linha.iloc[0]["tipo"]
        if tipo == "recesso":
            return "recesso"
        elif tipo == "feriado":
            return "final_semana"
    if data.weekday() >= 5:
        return "final_semana"
    return "noturno"


def prioridade_diurno(data):
    return PRIORIDADE_DIA_SEMANA.get(data.weekday(), 99)


def get_criticidade(nome):
    return criticidade_ferias.get(nome, 0)


# ========================
# SCORE PONDERADO NORMALIZADO
# ========================

def calcular_score_ponderado(agenda, datas, contador,
                              qtd_recesso, qtd_final, qtd_noturno, qtd_apoio,
                              pesos):
    p1, p2, p3, p4, p5 = pesos
    soma_pesos = p1 + p2 + p3 + p4 + p5
    if soma_pesos == 0:
        return 0.0, {}

    datas_list = list(datas)

    # Contagem de dias por tipo no mês
    contagem_tipo_dia = {}
    for d in datas_list:
        t = agenda[d]["tipo"]
        contagem_tipo_dia[t] = contagem_tipo_dia.get(t, 0) + 1

    dias_apoio = (
        contagem_tipo_dia.get("noturno", 0) +
        contagem_tipo_dia.get("final_semana", 0)
    )

    # ── C1: cobertura das listas especiais ─────────────────────────────────
    slots_esp_esperados = (
        qtd_recesso * contagem_tipo_dia.get("recesso", 0) +
        qtd_final   * contagem_tipo_dia.get("final_semana", 0) +
        qtd_noturno * contagem_tipo_dia.get("noturno", 0) +
        qtd_apoio   * dias_apoio
    )
    slots_esp_preenchidos = sum(
        len(agenda[d][t])
        for d in datas_list
        for t in ["Recesso", "Final de Semana", "Noturno", "Apoio"]
    )
    c1 = min(slots_esp_preenchidos / slots_esp_esperados, 1.0) if slots_esp_esperados > 0 else 1.0

    # ── C2: ausência de vagas vazias (especiais + diurnos) ──────────────────
    slots_diurnos_esperados = (
        sum(config_regioes[reg]["diurnos_por_dia"] for reg in regioes) *
        contagem_tipo_dia.get("noturno", 0)
    )
    slots_diurnos_preenchidos = sum(
        sum(len(nomes) for nomes in agenda[d]["Diurno"].values())
        for d in datas_list
    )
    slots_total_esperados   = slots_esp_esperados + slots_diurnos_esperados
    slots_total_preenchidos = slots_esp_preenchidos + slots_diurnos_preenchidos
    c2 = min(slots_total_preenchidos / slots_total_esperados, 1.0) if slots_total_esperados > 0 else 1.0

    # ── C3: equilíbrio por pessoa ───────────────────────────────────────────
    plantoes_pessoa = [
        sum(cnts.values()) for cnts in contador.values()
    ]
    if len(plantoes_pessoa) > 1 and sum(plantoes_pessoa) > 0:
        media  = sum(plantoes_pessoa) / len(plantoes_pessoa)
        desvio = math.sqrt(sum((x - media) ** 2 for x in plantoes_pessoa) / len(plantoes_pessoa))
        c3 = max(0.0, 1.0 - desvio / media)
    else:
        c3 = 1.0

    # ── C4: cobertura de datas críticas ────────────────────────────────────
    datas_criticas = [
        d for d in datas_list
        if d.weekday() in PRIORIDADE_DIA_SEMANA and agenda[d]["tipo"] == "noturno"
    ]
    max_criticos = sum(config_regioes[reg]["diurnos_por_dia"] for reg in regioes) * len(datas_criticas)
    diurnos_criticos = sum(
        sum(len(nomes) for nomes in agenda[d]["Diurno"].values())
        for d in datas_criticas
    )
    c4 = min(diurnos_criticos / max_criticos, 1.0) if max_criticos > 0 else 1.0

    # ── C5: equilíbrio de oficiais por dia ─────────────────────────────────
    plantoes_dia = [
        len(agenda[d]["Recesso"]) +
        len(agenda[d]["Final de Semana"]) +
        len(agenda[d]["Noturno"]) +
        len(agenda[d]["Apoio"]) +
        sum(len(nomes) for nomes in agenda[d]["Diurno"].values())
        for d in datas_list
    ]
    if len(plantoes_dia) > 1 and sum(plantoes_dia) > 0:
        media  = sum(plantoes_dia) / len(plantoes_dia)
        desvio = math.sqrt(sum((x - media) ** 2 for x in plantoes_dia) / len(plantoes_dia))
        c5 = max(0.0, 1.0 - desvio / media)
    else:
        c5 = 1.0

    # ── Score final ponderado ───────────────────────────────────────────────
    score = (p1 * c1 + p2 * c2 + p3 * c3 + p4 * c4 + p5 * c5) / soma_pesos

    detalhes = {
        "C1 Cobertura listas":   round(c1, 4),
        "C2 Ausência vazios":    round(c2, 4),
        "C3 Equilíbrio pessoa":  round(c3, 4),
        "C4 Datas críticas":     round(c4, 4),
        "C5 Equilíbrio dia":     round(c5, 4),
        "Score final":           round(score, 6),
    }
    return score, detalhes


# ========================
# PRÉ-GERAÇÃO DE LISTAS POR TIPO
# ========================

def gerar_lista_especial(tipo, todos_nomes, rng):
    """
    Retorna lista ordenada de candidatos para um tipo especial.
    Ordenação: data mais antiga → maior criticidade → ruído (só empates).
    """
    if tipo == "Apoio":
        candidatos = [
            {
                "NOME": nome,
                "ultima_data": pd.Timestamp("1900-01-01"),
                "criticidade": get_criticidade(nome),
                "ruido": rng.random(),
            }
            for nome in todos_nomes
        ]
    else:
        tipo_col = tipo.upper()
        sub = (
            df_long[df_long["tipo_plantao"] == tipo_col]
            .sort_values("ultima_data")
            .drop_duplicates(subset=["NOME"])
            .copy()
        )
        nomes_com_data = set(sub["NOME"].tolist())
        nomes_sem_data = [n for n in todos_nomes if n not in nomes_com_data]

        candidatos = []
        for _, row in sub.iterrows():
            candidatos.append({
                "NOME": row["NOME"],
                "ultima_data": row["ultima_data"],
                "criticidade": get_criticidade(row["NOME"]),
                "ruido": rng.random(),
            })
        for nome in nomes_sem_data:
            candidatos.append({
                "NOME": nome,
                "ultima_data": pd.Timestamp("2099-12-31"),
                "criticidade": get_criticidade(nome),
                "ruido": rng.random(),
            })

    candidatos.sort(key=lambda x: (x["ultima_data"], -x["criticidade"], x["ruido"]))
    return [c["NOME"] for c in candidatos]


# ========================
# GERADOR INTERNO
# ========================

def _gerar_escala_interna(ANO, MES, qtd_recesso, qtd_final, qtd_noturno, qtd_apoio, semente):

    rng = random.Random(semente)

    df_calendario = carregar_calendario(ANO)

    datas = pd.date_range(
        datetime(ANO, MES, 1),
        datetime(ANO, MES + 1, 1) - timedelta(days=1)
    )

    agenda = {
        data: {
            "tipo": tipo_do_dia(data, df_calendario),
            "Recesso": [],
            "Final de Semana": [],
            "Noturno": [],
            "Apoio": [],
            "Diurno": {}
        } for data in datas
    }

    todos_nomes = list(mapeamento_regioes.keys())
    nomes_por_regiao = {}
    for nome, reg in mapeamento_regioes.items():
        nomes_por_regiao.setdefault(reg, []).append(nome)

    contador = {
        nome: {"Diurno": 0, "Noturno": 0, "Final de Semana": 0, "Recesso": 0, "Apoio": 0}
        for nome in mapeamento_regioes
    }

    historico_diurno         = {nome: [] for nome in todos_nomes}
    historico_especiais      = {nome: [] for nome in todos_nomes}
    historico_noturno_semana = {nome: set() for nome in todos_nomes}
    historico_plantao        = {nome: [] for nome in todos_nomes}
    uso_mes                  = {nome: 0 for nome in todos_nomes}
    ja_fez_especial          = set()
    usados_dia               = {data: set() for data in datas}

    MAP_TIPO_DIA = {
        "Recesso":         "recesso",
        "Final de Semana": "final_semana",
        "Noturno":         "noturno",
        "Apoio":           None,  # filtrado por APOIO_TIPOS_VALIDOS
    }

    def elegivel_especial(nome, data, tipo):
        if nome in ja_fez_especial:
            return False
        if esta_de_ferias(nome, data):
            return False
        if nome in usados_dia[data]:
            return False
        if any(abs((data - d).days) < INTERVALO_MINIMO for d in historico_diurno[nome]):
            return False
        return True

    def preencher_tipo(tipo, quantidade):
        if quantidade == 0:
            return

        tipo_dia_alvo = MAP_TIPO_DIA[tipo]
        fila = gerar_lista_especial(tipo, todos_nomes, rng)

        for data in datas:
            # Filtro de tipo de dia
            if tipo == "Apoio":
                if agenda[data]["tipo"] not in APOIO_TIPOS_VALIDOS:
                    continue
            else:
                if agenda[data]["tipo"] != tipo_dia_alvo:
                    continue

            slots_restantes = quantidade - len(agenda[data][tipo])
            if slots_restantes <= 0:
                continue

            i = 0
            while i < len(fila) and slots_restantes > 0:
                nome = fila[i]
                if elegivel_especial(nome, data, tipo):
                    agenda[data][tipo].append(nome)
                    ja_fez_especial.add(nome)
                    usados_dia[data].add(nome)
                    contador[nome][tipo] += 1
                    historico_especiais[nome].append(data)
                    historico_plantao[nome].append(data)

                    if tipo == "Noturno":
                        semana = data.isocalendar()[1]
                        historico_noturno_semana[nome].add(semana)

                    fila.pop(i)
                    slots_restantes -= 1
                else:
                    i += 1

    preencher_tipo("Recesso",         qtd_recesso)
    preencher_tipo("Final de Semana", qtd_final)
    preencher_tipo("Noturno",         qtd_noturno)
    preencher_tipo("Apoio",           qtd_apoio)

    # ── Diurnos ─────────────────────────────────────────────────────────────
    datas_diurnas = sorted(
        [data for data in datas if agenda[data]["tipo"] == "noturno"],
        key=lambda d: (prioridade_diurno(d), d)
    )

    for data in datas_diurnas:
        semana = data.isocalendar()[1]

        for reg, nomes in nomes_por_regiao.items():
            qtd_reg = config_regioes[reg]["diurnos_por_dia"]
            max_mes = config_regioes[reg]["max_mes"]
            escolhidos = []

            candidatos_base = (
                df_long[df_long["NOME"].isin(nomes)]
                .sort_values("ultima_data")
                .drop_duplicates("NOME")
                .copy()
            )
            candidatos_base["tem_especial"]  = candidatos_base["NOME"].map(lambda n: 1 if n in ja_fez_especial else 0)
            candidatos_base["criticidade"]   = candidatos_base["NOME"].map(get_criticidade)
            candidatos_base["uso_mes_atual"] = candidatos_base["NOME"].map(uso_mes)
            candidatos_base["ruido"]         = [rng.random() for _ in range(len(candidatos_base))]
            candidatos_base = candidatos_base.sort_values(
                ["tem_especial", "uso_mes_atual", "ultima_data", "criticidade", "ruido"],
                ascending=[False, True, True, False, True]
            )

            for _, row in candidatos_base.iterrows():
                nome = row["NOME"]
                if len(escolhidos) >= qtd_reg:
                    break
                if esta_de_ferias(nome, data):
                    continue
                if uso_mes[nome] >= max_mes:
                    continue
                if any(abs((data - d).days) < INTERVALO_MINIMO for d in historico_especiais[nome]):
                    continue
                if any((data - d).days < 6 for d in historico_diurno[nome]):
                    continue
                if semana in historico_noturno_semana[nome]:
                    continue
                ultimos_14 = [d for d in historico_plantao[nome] if (data - d).days <= 14]
                if len(ultimos_14) >= 3:
                    continue

                escolhidos.append(nome)
                historico_diurno[nome].append(data)
                historico_plantao[nome].append(data)
                uso_mes[nome] += 1
                contador[nome]["Diurno"] += 1

            agenda[data]["Diurno"][reg] = escolhidos

    return agenda, datas, contador


def _formatar_resultado(agenda, datas, contador):
    df_resultado = pd.DataFrame([
        {
            "Data": d.strftime("%d/%m/%Y"),
            "Dia": d.strftime("%A"),
            "Tipo": agenda[d]["tipo"],
            "Recesso": ", ".join(agenda[d]["Recesso"]),
            "Final de Semana": ", ".join(agenda[d]["Final de Semana"]),
            "Noturno": ", ".join(agenda[d]["Noturno"]),
            "Apoio": ", ".join(agenda[d]["Apoio"]),
            "Diurno": str(agenda[d]["Diurno"])
        } for d in datas
    ])
    df_contador = pd.DataFrame.from_dict(contador, orient="index")
    df_contador["Total"] = df_contador.sum(axis=1)
    df_contador["Dias de Férias (Criticidade)"] = df_contador.index.map(get_criticidade)
    return df_resultado, df_contador.reset_index().rename(columns={"index": "Nome"})


# ========================
# VISUALIZAÇÃO DAS LISTAS PRÉ-GERADAS
# ========================

def gerar_df_listas(semente):
    rng = random.Random(semente)
    todos_nomes = list(mapeamento_regioes.keys())
    listas = {}
    max_len = 0
    for tipo in ["Recesso", "Final de Semana", "Noturno", "Apoio"]:
        lista = gerar_lista_especial(tipo, todos_nomes, rng)
        listas[tipo] = lista
        max_len = max(max_len, len(lista))

    rows = []
    for i in range(max_len):
        row = {"#": i + 1}
        for tipo in ["Recesso", "Final de Semana", "Noturno", "Apoio"]:
            row[tipo] = listas[tipo][i] if i < len(listas[tipo]) else ""
        rows.append(row)
    return pd.DataFrame(rows)


# ========================
# OTIMIZADOR
# ========================

def gerar_melhor_escala(ANO, MES, qtd_recesso, qtd_final, qtd_noturno, qtd_apoio,
                         n_iteracoes, pesos):
    melhor_agenda    = None
    melhor_datas     = None
    melhor_contador  = None
    melhor_score     = -1.0
    melhor_detalhes  = {}
    melhor_semente   = 0
    historico_scores = []

    barra = st.progress(0, text="Calculando iterações...")

    for i in range(n_iteracoes):
        semente = i
        agenda, datas, contador = _gerar_escala_interna(
            ANO, MES, qtd_recesso, qtd_final, qtd_noturno, qtd_apoio, semente
        )
        score, detalhes = calcular_score_ponderado(
            agenda, datas, contador,
            qtd_recesso, qtd_final, qtd_noturno, qtd_apoio,
            pesos
        )

        historico_scores.append({
            "Iteração":             i + 1,
            "Semente":              semente,
            "Score final":          round(score, 6),
            "C1 Cobertura listas":  detalhes.get("C1 Cobertura listas", 0),
            "C2 Ausência vazios":   detalhes.get("C2 Ausência vazios", 0),
            "C3 Equilíbrio pessoa": detalhes.get("C3 Equilíbrio pessoa", 0),
            "C4 Datas críticas":    detalhes.get("C4 Datas críticas", 0),
            "C5 Equilíbrio dia":    detalhes.get("C5 Equilíbrio dia", 0),
        })

        if score > melhor_score:
            melhor_score    = score
            melhor_detalhes = detalhes
            melhor_agenda   = agenda
            melhor_datas    = datas
            melhor_contador = contador
            melhor_semente  = semente

        barra.progress(
            (i + 1) / n_iteracoes,
            text=f"Iteração {i + 1}/{n_iteracoes} | Melhor score: {melhor_score:.4f}"
        )

    barra.empty()
    df_resultado, df_contador = _formatar_resultado(melhor_agenda, melhor_datas, melhor_contador)
    df_scores = pd.DataFrame(historico_scores)
    return (df_resultado, df_contador, melhor_score, melhor_detalhes,
            df_scores, melhor_agenda, melhor_datas, melhor_semente)


# ========================
# EXPORTAÇÃO EXCEL
# ========================

DIAS_PT = {
    "Monday": "segunda-feira", "Tuesday": "terça-feira",
    "Wednesday": "quarta-feira", "Thursday": "quinta-feira",
    "Friday": "sexta-feira", "Saturday": "sábado", "Sunday": "domingo",
}

COR = {
    "cab_data":   "4472C4",
    "cab_dia":    "D9E1F2",
    "reg_label":  "2F4F4F",
    "diurno_a":   "FFF2CC",
    "diurno_b":   "FFFFFF",
    "fds_bg":     "FFF0F0",
    "sep":        "F2F2F2",
    "ps_label":   "1F3864",
    "ps_noturno": "BDD7EE",
    "ps_fds":     "FCE4D6",
    "ps_recesso": "E2EFDA",
    "ap_label":   "595959",
    "ap_bg":      "F2F2F2",
}

def _f(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _b(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def gerar_excel(agenda, datas, mapeamento_regioes, config_regioes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Escala"

    regioes_ord = sorted(set(mapeamento_regioes.values()))
    datas_list  = list(datas)
    n_datas     = len(datas_list)

    slots_por_regiao = {
        reg: max(config_regioes[reg]["diurnos_por_dia"], 1)
        for reg in regioes_ord
    }

    # Linha 1 — datas
    ws.cell(row=1, column=1, value="").fill = _f(COR["reg_label"])
    ws.cell(row=1, column=1).border = _b()
    for ci, data in enumerate(datas_list, start=2):
        c = ws.cell(row=1, column=ci, value=data.strftime("%d/%m/%y"))
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.fill      = _f(COR["cab_data"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _b()

    # Linha 2 — dia da semana
    ws.cell(row=2, column=1, value="").fill = _f(COR["reg_label"])
    ws.cell(row=2, column=1).border = _b()
    for ci, data in enumerate(datas_list, start=2):
        dia = DIAS_PT.get(data.strftime("%A"), data.strftime("%A"))
        c = ws.cell(row=2, column=ci, value=dia)
        c.font      = Font(size=9, color="444444", name="Arial")
        c.fill      = _f(COR["cab_dia"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _b()

    cur = 3

    # Blocos de diurno por região
    for ri, reg in enumerate(regioes_ord):
        n_slots    = slots_por_regiao[reg]
        cor_diurno = COR["diurno_a"] if ri % 2 == 0 else COR["diurno_b"]

        for slot in range(n_slots):
            row = cur + slot
            lc = ws.cell(row=row, column=1)
            lc.value     = reg if slot == 0 else ""
            lc.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            lc.fill      = _f(COR["reg_label"])
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border    = _b()

            for ci, data in enumerate(datas_list, start=2):
                nomes    = agenda[data]["Diurno"].get(reg, [])
                valor    = nomes[slot] if slot < len(nomes) else ""
                tipo_dia = agenda[data]["tipo"]
                bg = COR["fds_bg"] if tipo_dia in ("final_semana", "recesso") else cor_diurno
                c = ws.cell(row=row, column=ci, value=valor)
                c.fill      = _f(bg)
                c.font      = Font(size=9, bold=bool(valor), name="Arial")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = _b()

        cur += n_slots

        for ci in range(1, n_datas + 2):
            c = ws.cell(row=cur, column=ci, value="")
            c.fill   = _f(COR["sep"])
            c.border = _b()
        cur += 1

    # Linha PS
    max_ps = max(
        max(
            len(agenda[d]["Recesso"]) +
            len(agenda[d]["Final de Semana"]) +
            len(agenda[d]["Noturno"])
            for d in datas_list
        ),
        1
    )

    for slot in range(max_ps):
        row = cur + slot
        lc = ws.cell(row=row, column=1)
        lc.value     = "PS" if slot == 0 else ""
        lc.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        lc.fill      = _f(COR["ps_label"])
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _b()

        for ci, data in enumerate(datas_list, start=2):
            itens = (
                [(n, "Recesso")         for n in agenda[data]["Recesso"]] +
                [(n, "Final de Semana") for n in agenda[data]["Final de Semana"]] +
                [(n, "Noturno")         for n in agenda[data]["Noturno"]]
            )
            tipo_dia = agenda[data]["tipo"]

            if slot < len(itens):
                nome, tipo_esp = itens[slot]
                valor = f"{nome} {tipo_esp}"
                bg = (
                    COR["ps_recesso"] if tipo_esp == "Recesso"
                    else COR["ps_fds"] if tipo_esp == "Final de Semana"
                    else COR["ps_noturno"]
                )
            else:
                valor = ""
                bg = COR["ps_fds"] if tipo_dia in ("final_semana", "recesso") else COR["ps_noturno"]

            c = ws.cell(row=row, column=ci, value=valor)
            c.fill      = _f(bg)
            c.font      = Font(size=9, bold=bool(valor), name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _b()

    cur += max_ps

    # Linha AP
    max_ap = max(
        max((len(agenda[d]["Apoio"]) for d in datas_list), default=0),
        1
    )

    for slot in range(max_ap):
        row = cur + slot
        lc = ws.cell(row=row, column=1)
        lc.value     = "AP" if slot == 0 else ""
        lc.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        lc.fill      = _f(COR["ap_label"])
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _b()

        for ci, data in enumerate(datas_list, start=2):
            apoios = agenda[data]["Apoio"]
            valor  = apoios[slot] if slot < len(apoios) else ""
            c = ws.cell(row=row, column=ci, value=valor)
            c.fill      = _f(COR["ap_bg"])
            c.font      = Font(size=9, color="666666", name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _b()

    ws.column_dimensions["A"].width = 7
    for ci in range(2, n_datas + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ========================
# UI
# ========================

with st.expander("👁️ Prévia das listas de candidatos por tipo especial (semente 0)"):
    st.caption(
        "Estas são as filas base (semente 0) que alimentam o algoritmo. "
        "Cada iteração aplica um ruído diferente nos empates, gerando variações. "
        "Os nomes serão consumidos em ordem, com rollback se não puderem ser encaixados. "
        "⚠️ O Apoio só ocorre em dias de plantão Noturno e Final de Semana — nunca em Recesso."
    )
    df_listas = gerar_df_listas(semente=0)
    st.dataframe(df_listas, use_container_width=True, height=400)

if st.button("🚀 Gerar Escala"):

    pesos = (
        peso_cobertura_listas,
        peso_vagas_vazias,
        peso_equil_pessoa,
        peso_datas_criticas,
        peso_equil_dia,
    )

    (df, df_contador, melhor_score, melhor_detalhes,
     df_scores, melhor_agenda, melhor_datas, melhor_semente) = gerar_melhor_escala(
        ANO, MES, qtd_recesso, qtd_final, qtd_noturno, qtd_apoio,
        int(n_iteracoes), pesos
    )

    # ── Painel de pontuação ─────────────────────────────────────────────────
    st.subheader("🏆 Melhor Escala Encontrada")

    col_score, col_semente = st.columns([3, 1])
    with col_score:
        st.metric(
            "Score final ponderado", f"{melhor_score:.4f}",
            help="Score normalizado entre 0 e 1. Quanto mais próximo de 1, melhor."
        )
    with col_semente:
        st.metric("Melhor semente", melhor_semente)

    st.caption("Detalhamento dos componentes — cada valor entre 0 (ruim) e 1 (perfeito):")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("C1 Cobertura listas",  f"{melhor_detalhes.get('C1 Cobertura listas',  0):.3f}",
              help="Slots especiais preenchidos / slots esperados")
    c2.metric("C2 Ausência vazios",   f"{melhor_detalhes.get('C2 Ausência vazios',   0):.3f}",
              help="Total de slots preenchidos / total esperado")
    c3.metric("C3 Equilíbrio pessoa", f"{melhor_detalhes.get('C3 Equilíbrio pessoa', 0):.3f}",
              help="1 − coef. de variação dos plantões por pessoa")
    c4.metric("C4 Datas críticas",    f"{melhor_detalhes.get('C4 Datas críticas',    0):.3f}",
              help="Diurnos em dias prioritários / máximo possível")
    c5.metric("C5 Equilíbrio dia",    f"{melhor_detalhes.get('C5 Equilíbrio dia',    0):.3f}",
              help="1 − coef. de variação dos plantões por dia")

    # ── Listas da melhor iteração ───────────────────────────────────────────
    with st.expander("📋 Listas de candidatos da melhor iteração"):
        st.caption(
            f"Listas geradas com semente {melhor_semente} — "
            "estas foram as filas que produziram a melhor escala."
        )
        df_listas_melhor = gerar_df_listas(semente=melhor_semente)
        st.dataframe(df_listas_melhor, use_container_width=True, height=400)

    # ── Comparativo de iterações ────────────────────────────────────────────
    with st.expander("📊 Comparativo de todas as iterações"):
        def destacar_melhor(df):
            scores     = df["Score final"].tolist()
            idx_melhor = scores.index(max(scores))
            estilos = [
                ["background-color: #d4edda; font-weight: bold"
                 if i == idx_melhor else "" for _ in df.columns]
                for i in range(len(df))
            ]
            return pd.DataFrame(estilos, index=df.index, columns=df.columns)

        st.dataframe(
            df_scores.style.apply(destacar_melhor, axis=None),
            use_container_width=True
        )

    # ── Escala e contadores ─────────────────────────────────────────────────
    st.subheader("📅 Escala")
    st.dataframe(df, use_container_width=True)

    st.subheader("📊 Contador de Plantões")
    st.dataframe(df_contador, use_container_width=True)

    # ── Criticidade ─────────────────────────────────────────────────────────
    st.subheader("🚨 Índice de Criticidade por Pessoa")
    df_crit = pd.DataFrame([
        {"Nome": nome, "Dias de Férias": dias}
        for nome, dias in sorted(
            criticidade_ferias.items(), key=lambda x: x[1], reverse=True
        )
        if dias > 0
    ])
    if not df_crit.empty:
        st.dataframe(df_crit, use_container_width=True)
    else:
        st.info("Nenhuma pessoa com férias marcadas no mês.")

    # ── Exportar Excel ──────────────────────────────────────────────────────
    st.subheader("📥 Exportar Planilha")
    excel_buf = gerar_excel(melhor_agenda, melhor_datas, mapeamento_regioes, config_regioes)
    st.download_button(
        label="⬇️ Baixar planilha Excel (.xlsx)",
        data=excel_buf,
        file_name=f"escala_{ANO}_{MES:02d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
